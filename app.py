import base64
import os
import re
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from io import BytesIO
from urllib.parse import quote_plus, urljoin
from typing import Optional, List, Set, Tuple, Dict, Any
from copy import copy as shallow_copy
from threading import Lock

import requests
from bs4 import BeautifulSoup, Tag
from flask import Flask, render_template, url_for, request, jsonify
from openpyxl import load_workbook
from openpyxl.cell import Cell

# Playwright imports (optional, теперь не обязательно для IPRbooks)
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
    PLAYWRIGHT_AVAILABLE = True
    PLAYWRIGHT_LOCK = Lock()
except ImportError:
    print("[INFO] Playwright не установлен. Используется requests для IPRbooks поиска.")
    sync_playwright = None
    PlaywrightTimeoutError = Exception
    PLAYWRIGHT_AVAILABLE = False
    PLAYWRIGHT_LOCK = None

BANNED_SUBSTRINGS = (
    "иностран",
    "физкульт", 
    "физическая культура",
    "физкультура",
    "история казахстана",
    "информационно",
    "коммуникативн",
    "философия",
    "шетел тілі",
    "қазақ",
    "орыс",
    "тілі",
    "қазақстан тарихы",
    "ақпараттық",
    "коммуникациялық",
    "технологиялар",
    "дене шынықтыру",
    "жббп",
    "бп жк",
    "бп тк", 
    "бп циклі",
    "кп жк",
    "кп тк",
    "кп циклі",
    "циклі бойынша",
    "по циклу",
    "оод ок",
    "оод вк",
    "оод",
    "бд вк",
    "бд кв",
    "бд",
    "пд вк", 
    "пд кв",
    "пд",
    "эирм",
    "қорытынды аттестаттау",
    "итоговая аттестация",
    "final certification",
    "академиялық департамент",
    "директор академического департамента",
    "нужно так",
)
SUBJECT_EXCLUDE_SUBSTRINGS = (
    "учебная нагрузка",
    "учебн",
    "семинар",
    "практичес",
    "лаборатор",
    "самостоятель",
    "руководител",
    "организац",
    "декан",
    "заместитель",
    "профессор",
    "кандидат",
    "научно",
    "деятельност",
    "работ",
)
MAX_WORKERS = 8
PLAYWRIGHT_LOCK = Lock() if PLAYWRIGHT_AVAILABLE else None

KNOWN_RESOURCE_RULES = [
    {
        "contains": ("физ", "культур"),
        "resources": [
            {
                "title": "Веденина О.А. Самостоятельные занятия физической культурой : учебное пособие для СПО / Веденина О.А.. — Саратов, Москва : Профобразование, Ай Пи Ар Медиа, 2025.",
                "url": "https://www.iprbookshop.ru/147933.html",
                "status": "success",
                "note": "Учебное пособие IPRbooks",
            },
            {
                "title": "Гришина, Ю.И. Физическая культура студента: учеб. пособие. — Ростов н/Д: Феникс, 2019. — 283 c. — (Высшее образование). — ISBN 978-5-222-31286-5.",
                "url": "https://rmebrk.kz/book/1177773",
                "status": "warning",
                "note": "Доступ на rmebrk.kz",
            },
        ],
    },
]

# Cookies для авторизации на IPRbooks
# ВАЖНО: Если поиск перестанет работать на Vercel, возможно нужно обновить cookies
# Получить актуальные cookies можно через DevTools браузера после авторизации на сайте
IPRBOOKSHOP_COOKIES = {
    ".iprbookshop.ru": {
        "_ym_d": "1762763770",
        "_ym_uid": "1762763770129598589",
    },
    "www.iprbookshop.ru": {
        "privacy-policy": "1",
        "read-vzu": "1",
        "SN4f61b1c8b1bd0": "4sqmd0q49i06iilh542hr77d57",
    },
    ".www.iprbookshop.ru": {
        "IPRSMARTLogin": "89eafbcebab37c937a067ad7671a26b9%7C0a31eb43929401ec874925b9091c4ba6",
    },
}

app = Flask(__name__)


@app.route("/")
def home():
    return render_template("index.html", current_year=datetime.now().year)


@app.route("/search")
def search():
    return render_template("search.html", current_year=datetime.now().year)


def extract_subjects_from_up33(up_file_stream) -> List[str]:
    wb = load_workbook(up_file_stream, data_only=True)
    ws = wb.active
    subjects: List[str] = []
    seen: Set[str] = set()
    for r in range(2, ws.max_row + 1):
        c_val = ws.cell(row=r, column=3).value
        d_val = ws.cell(row=r, column=4).value
        c_text = str(c_val).strip() if c_val is not None else ""
        d_text = str(d_val).strip() if d_val is not None else ""
        if not c_text and not d_text:
            continue
        header_tokens = {"наимен", "предмет", "дисциплин"}
        combined_text = " ".join([c_text.lower(), d_text.lower()]).strip()
        if any(token in combined_text for token in header_tokens):
            continue
        merged = " / ".join([text for text in (c_text, d_text) if text])
        
        # Извлекаем только русскую часть (после "/")
        if " / " in merged:
            parts = merged.split(" / ")
            # Берем последнюю часть (обычно русская)
            russian_part = parts[-1].strip()
            if russian_part:
                merged = russian_part
        
        merged_lower = merged.lower()
        if any(bad in merged_lower for bad in SUBJECT_EXCLUDE_SUBSTRINGS):
            continue
        if any(bad in merged_lower for bad in BANNED_SUBSTRINGS):
            continue
        if len(merged_lower) > 160 or len(merged_lower.split()) > 20:
            continue
        norm = merged.strip().lower()
        if merged and norm not in seen:
            seen.add(norm)
            subjects.append(merged)
    return subjects


def find_next_row(ws, start_row: int) -> int:
    r = start_row
    while True:
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        if not any([a, b, c, d]):
            return r
        r += 1


def compute_next_number(ws, start_row: int) -> int:
    max_num = 0
    for r in range(start_row, start_row + 2000):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, (int, float)):
            try:
                iv = int(v)
                if iv > max_num:
                    max_num = iv
            except Exception:
                pass
        elif isinstance(v, str) and v.strip().isdigit():
            iv = int(v.strip())
            if iv > max_num:
                max_num = iv
    return max_num + 1 if max_num >= 0 else 1


def search_rmebrk_results(subject: str, max_results: int = 10) -> List[Dict[str, Any]]:
    """Поиск на РМЭБ (Республиканская Межвузовская Электронная Библиотека)"""
    base_url = "https://rmebrk.kz/"
    results = []

    # Определяем, находимся ли мы на Vercel (serverless)
    is_vercel = os.getenv('VERCEL') == '1' or 'vercel' in os.getenv('VERCEL_URL', '').lower()

    try:
        # Создаем сессию с правильными заголовками
        session = requests.Session()

        # Для Vercel используем более простой User-Agent
        if is_vercel:
            session.headers.update({
                'User-Agent': 'Mozilla/5.0 (compatible; UniBiblio/1.0)',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'ru,en-US;q=0.9,en;q=0.8',
            })
        else:
            session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
                'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
                'Accept-Encoding': 'gzip, deflate, br',
                'DNT': '1',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'none',
                'Cache-Control': 'max-age=0',
            })

        # Короткие таймауты для Vercel
        timeout = 8 if is_vercel else 15

        # Сначала загружаем главную страницу для получения сессии
        print(f"[DEBUG] RMЭБ: загрузка главной страницы (Vercel: {is_vercel})")
        try:
            main_response = session.get(base_url, timeout=timeout)
        except Exception as e:
            print(f"[DEBUG] RMЭБ: ошибка загрузки главной страницы: {e}, пропускаем")
            return []

        if main_response.status_code != 200:
            print(f"[DEBUG] RMЭБ: главная страница вернула {main_response.status_code}, пропускаем")
            return []

        # Небольшая задержка для Vercel
        if is_vercel:
            import time
            time.sleep(0.5)

        # Прямой подход: используем поисковый URL с параметром 'search'
        # На основе описания пользователя: поле name="search", id="_searchinput"
        search_data = {
            'search': subject.strip()
        }

        # Убираем лишний слеш в конце base_url
        base_url_clean = base_url.rstrip('/')
        search_url_post = f"{base_url_clean}/search"
        
        print(f"[DEBUG] RMЭБ: пробуем прямой POST на {search_url_post} с параметром 'search'")

        try:
            # Сначала пробуем POST на /search
            search_response = session.post(search_url_post, data=search_data, timeout=timeout)

            if search_response.status_code != 200 or len(search_response.text) < 1000:
                print(f"[DEBUG] RMЭБ: POST вернул {search_response.status_code}, длина: {len(search_response.text)}, пробуем GET")
                # Если POST не сработал, пробуем GET с параметрами в URL
                search_url_get = f"{base_url_clean}/search?search={quote_plus(subject.strip())}"
                print(f"[DEBUG] RMЭБ: пробуем GET запрос: {search_url_get}")
                search_response = session.get(search_url_get, timeout=timeout)

            if search_response.status_code != 200 or len(search_response.text) < 1000:
                print(f"[DEBUG] RMЭБ: поиск не удался, пробуем анализ главной страницы")
                # Анализируем главную страницу для поиска поля
                main_soup = BeautifulSoup(main_response.text, 'html.parser')

                # Логируем все input поля на странице
                all_inputs = main_soup.find_all('input')
                print(f"[DEBUG] RMЭБ: найдено input полей на странице: {len(all_inputs)}")
                for i, inp in enumerate(all_inputs[:15]):  # Показываем первые 15
                    inp_type = inp.get('type', 'text')
                    print(f"[DEBUG] RMЭБ: input {i+1}: name={inp.get('name')}, id={inp.get('id')}, type={inp_type}, placeholder={inp.get('placeholder')}")

                # Ищем конкретное поле _searchinput или поле с name="search"
                search_input = main_soup.find('input', {'id': '_searchinput'})
                if not search_input:
                    search_input = main_soup.find('input', {'name': 'search'})

                if search_input:
                    print(f"[DEBUG] RMЭБ: найдено поле поиска: name={search_input.get('name')}, id={search_input.get('id')}")
                    # Пробуем POST с правильным параметром
                    param_name = search_input.get('name') or 'search'
                    search_data = {param_name: subject.strip()}
                    search_response = session.post(search_url_post, data=search_data, timeout=timeout)
                    print(f"[DEBUG] RMЭБ: повторный POST с параметром '{param_name}'")
                else:
                    print(f"[DEBUG] RMЭБ: поле поиска не найдено, пропускаем")
                    return []

        except Exception as e:
            print(f"[DEBUG] RMЭБ: ошибка поиска: {e}, пропускаем")
            return []

        if search_response.status_code != 200:
            print(f"[DEBUG] RMЭБ: поиск вернул {search_response.status_code}, пропускаем")
            return []

        search_response.encoding = 'utf-8'
        print(f"[DEBUG] RMЭБ: страница результатов получена, длина: {len(search_response.text)} символов")
        print(f"[DEBUG] RMЭБ: URL результатов: {search_response.url}")

        # Ищем AJAX endpoint в JavaScript коде страницы
        # Результаты могут загружаться динамически через AJAX
        ajax_endpoint = None
        ajax_params = {}
        
        # Ищем в JavaScript коде упоминания AJAX запросов
        if '/test/listinlist' in search_response.text or '/api/' in search_response.text or '/ajax/' in search_response.text:
            # Пробуем найти endpoint в тексте страницы
            # Ищем паттерны типа /test/listinlist, /api/search, /ajax/search и т.д.
            endpoint_patterns = [
                r'["\'](/test/[^"\']+)["\']',
                r'["\'](/api/[^"\']+)["\']',
                r'["\'](/ajax/[^"\']+)["\']',
                r'url\s*[:=]\s*["\']([^"\']+)["\']',
                r'action\s*[:=]\s*["\']([^"\']+)["\']',
            ]
            
            for pattern in endpoint_patterns:
                matches = re.findall(pattern, search_response.text)
                if matches:
                    potential_endpoint = matches[0]
                    if any(word in potential_endpoint.lower() for word in ['search', 'list', 'book', 'publication', 'resource']):
                        ajax_endpoint = potential_endpoint
                        print(f"[DEBUG] RMЭБ: найден потенциальный AJAX endpoint: {ajax_endpoint}")
                        break
        
        # Используем переменную для HTML контента (может быть заменена AJAX ответом)
        html_content = search_response.text
        
        # Пробуем известный AJAX endpoint /test/listinlist
        ajax_tried = False
        if not ajax_endpoint:
            ajax_endpoint = '/test/listinlist'
            print(f"[DEBUG] RMЭБ: используем известный AJAX endpoint: {ajax_endpoint}")
        
        # Если нашли AJAX endpoint, пробуем использовать его
        if ajax_endpoint:
            try:
                ajax_url = urljoin(base_url_clean, ajax_endpoint)
                print(f"[DEBUG] RMЭБ: пробуем AJAX запрос на {ajax_url}")
                
                # Пробуем разные варианты параметров (сначала form-data, потом JSON)
                ajax_data_variants = [
                    # Form data варианты
                    {'keyword': subject.strip(), 'secondSearchVar': '', 'pagination': 1, 'orderby': ''},
                    {'search': subject.strip(), 'page': 1},
                    {'q': subject.strip(), 'page': 1},
                    {'query': subject.strip()},
                ]
                
                for ajax_data in ajax_data_variants:
                    try:
                        # Сначала пробуем POST с form-data
                        ajax_response = session.post(ajax_url, data=ajax_data, timeout=timeout, headers={
                            'X-Requested-With': 'XMLHttpRequest',
                            'Referer': search_response.url,
                            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                        })
                        if ajax_response.status_code == 200:
                            ajax_html = ajax_response.text
                            print(f"[DEBUG] RMЭБ: AJAX вернул ответ, длина: {len(ajax_html)}")
                            # Проверяем, есть ли в ответе элементы результатов
                            if len(ajax_html) > 1000 and ('list-group-item' in ajax_html or '/book/' in ajax_html):
                                html_content = ajax_html  # Используем AJAX ответ
                                ajax_tried = True
                                print(f"[DEBUG] RMЭБ: используем AJAX ответ для парсинга (найдены результаты)")
                                break
                    except Exception as e:
                        print(f"[DEBUG] RMЭБ: ошибка AJAX запроса (form-data): {e}")
                        continue
                
                # Если form-data не сработало, пробуем JSON
                if not ajax_tried:
                    for ajax_data in ajax_data_variants:
                        try:
                            ajax_response = session.post(ajax_url, json=ajax_data, timeout=timeout, headers={
                                'X-Requested-With': 'XMLHttpRequest',
                                'Content-Type': 'application/json',
                                'Accept': 'application/json',
                                'Referer': search_response.url,
                            })
                            if ajax_response.status_code == 200:
                                try:
                                    ajax_json = ajax_response.json()
                                    print(f"[DEBUG] RMЭБ: AJAX JSON ответ получен")
                                    # Если это JSON с данными, обрабатываем
                                    if isinstance(ajax_json, dict):
                                        print(f"[DEBUG] RMЭБ: JSON ключи: {list(ajax_json.keys())}")
                                    print(f"[DEBUG] RMЭБ: AJAX ответ: {str(ajax_json)[:500]}")
                                except:
                                    # Не JSON, возможно HTML
                                    ajax_html = ajax_response.text
                                    print(f"[DEBUG] RMЭБ: AJAX вернул HTML, длина: {len(ajax_html)}")
                                    if len(ajax_html) > 1000 and ('list-group-item' in ajax_html or '/book/' in ajax_html):
                                        html_content = ajax_html
                                        ajax_tried = True
                                        break
                        except Exception as e:
                            print(f"[DEBUG] RMЭБ: ошибка AJAX запроса (JSON): {e}")
                            continue
            except Exception as e:
                print(f"[DEBUG] RMЭБ: ошибка при попытке AJAX: {e}")

        # Парсим страницу результатов (используем html_content, который может быть из AJAX)
        result_soup = BeautifulSoup(html_content, 'html.parser')

        # Логируем структуру страницы для диагностики
        print(f"[DEBUG] RMЭБ: анализируем структуру страницы результатов...")
        print(f"[DEBUG] RMЭБ: длина HTML контента: {len(html_content)} символов")
        print(f"[DEBUG] RMЭБ: используется AJAX ответ: {ajax_tried}")
        
        # Проверяем, есть ли в HTML элементы результатов
        if 'list-group-item' in html_content:
            print(f"[DEBUG] RMЭБ: найдены элементы list-group-item в HTML")
        if '/book/' in html_content:
            print(f"[DEBUG] RMЭБ: найдены ссылки /book/ в HTML")
        
        # Ищем все div с классами col-md-*
        all_col_divs = result_soup.find_all('div', class_=lambda x: x and ('col-md' in str(x) or 'col-xs' in str(x) or 'col-sm' in str(x)))
        print(f"[DEBUG] RMЭБ: найдено div с классами col-md/col-xs/col-sm: {len(all_col_divs)}")
        
        # Сначала логируем все ссылки на странице для диагностики
        all_page_links = result_soup.find_all('a', href=True)
        print(f"[DEBUG] RMЭБ: найдено всего ссылок на странице: {len(all_page_links)}")
        for i, link in enumerate(all_page_links[:20]):  # Показываем первые 20
            href = link.get('href', '')
            text = link.get_text(strip=True)[:50]
            classes = link.get('class', [])
            print(f"[DEBUG] RMЭБ: ссылка {i+1}: href={href[:100]}, text={text}, class={classes}")
        
        # Ищем контейнер с результатами (как описано в инструкции пользователя)
        results_container = result_soup.find('div', {'class': 'col-md-12 col-xs-12 col-sm-12'})
        if not results_container:
            # Пробуем другие варианты селекторов
            results_container = result_soup.find('div', class_=lambda x: x and 'col-md-12' in str(x) and 'col-xs-12' in str(x))
            if not results_container:
                # Ищем любой div с col-md-12
                results_container = result_soup.find('div', class_=lambda x: x and 'col-md-12' in str(x))
            if not results_container:
                # Ищем контейнеры с результатами по другим признакам
                # Пробуем найти список результатов (ul, ol) или контейнеры с книгами
                results_container = result_soup.find('ul', class_=lambda x: x and ('list' in str(x).lower() or 'result' in str(x).lower()))
                if not results_container:
                    results_container = result_soup.find('div', class_=lambda x: x and ('result' in str(x).lower() or 'book' in str(x).lower() or 'item' in str(x).lower()))
        
        if results_container:
            print(f"[DEBUG] RMЭБ: найден контейнер результатов: {results_container.name}, class={results_container.get('class')}")
            # Логируем содержимое контейнера
            container_text = results_container.get_text(strip=True)[:500]
            print(f"[DEBUG] RMЭБ: текст контейнера (первые 500 символов): {container_text}")
            container_html = str(results_container)[:1000]
            print(f"[DEBUG] RMЭБ: HTML контейнера (первые 1000 символов): {container_html}")
        else:
            print(f"[DEBUG] RMЭБ: контейнер результатов не найден")
            # Логируем первые 2000 символов HTML для анализа
            print(f"[DEBUG] RMЭБ: превью HTML (первые 2000 символов):")
            print(search_response.text[:2000])
        
        # Если контейнер найден но пустой, или не найден - используем весь документ
        if not results_container or len(results_container.find_all('a', href=True)) == 0:
            print(f"[DEBUG] RMЭБ: контейнер пустой или не найден, используем весь документ")
            results_container = result_soup

        # Ищем элементы результатов поиска - <li class="list-group-item">
        result_items = results_container.find_all('li', class_=lambda x: x and 'list-group-item' in str(x))
        print(f"[DEBUG] RMЭБ: найдено элементов результатов (list-group-item): {len(result_items)}")
        
        # Ищем все ссылки на книги в результатах
        all_links = results_container.find_all('a', href=True)
        print(f"[DEBUG] RMЭБ: найдено {len(all_links)} ссылок в контейнере результатов")

        # Также ищем элементы с data-link или onclick (могут быть ссылками на книги)
        data_link_elements = results_container.find_all(attrs={'data-link': True})
        onclick_elements = results_container.find_all(attrs={'onclick': True})
        # Ищем элементы с data-id (могут быть ссылками на книги)
        data_id_elements = results_container.find_all(attrs={'data-id': True})
        print(f"[DEBUG] RMЭБ: найдено элементов с data-link: {len(data_link_elements)}, с onclick: {len(onclick_elements)}, с data-id: {len(data_id_elements)}")
        
        # Обрабатываем элементы результатов - ищем ссылки на книги внутри них
        print(f"[DEBUG] RMЭБ: обрабатываем {len(result_items)} элементов результатов")
        for i, item in enumerate(result_items):
            print(f"[DEBUG] RMЭБ: элемент результата {i+1}:")
            
            # Ищем все ссылки внутри элемента результата (обычные <a> теги)
            item_links = item.find_all('a', href=True)
            print(f"[DEBUG] RMЭБ: найдено {len(item_links)} ссылок <a> в элементе {i+1}")
            
            # Ищем ссылку "Просмотр" с /book/{id} в обычных ссылках
            view_link = None
            for link in item_links:
                href = link.get('href', '')
                link_text = link.get_text(strip=True).lower()
                print(f"[DEBUG] RMЭБ: ссылка <a> в элементе {i+1}: href={href[:100]}, text={link_text[:50]}")
                if re.match(r'^/book/\d+', href):
                    view_link = link
                    print(f"[DEBUG] RMЭБ: найдена ссылка на книгу в <a>: {href}")
                    break
            
            # Также ищем элементы с data-link атрибутом (например, <li class="nopublic_book" data-link="/book/123">)
            if not view_link:
                data_link_elems = item.find_all(attrs={'data-link': True})
                print(f"[DEBUG] RMЭБ: найдено {len(data_link_elems)} элементов с data-link в элементе {i+1}")
                for data_link_elem in data_link_elems:
                    data_link = data_link_elem.get('data-link', '')
                    print(f"[DEBUG] RMЭБ: элемент с data-link: {data_link[:100]}")
                    if re.match(r'^/book/\d+', data_link):
                        # Создаем искусственную ссылку из data-link
                        class DataLinkLink:
                            def __init__(self, data_link, elem, item):
                                self.data_link = data_link
                                self.elem = elem
                                self.item = item
                                self.href = data_link
                            def get(self, attr):
                                if attr == 'href':
                                    return self.href
                                return self.elem.get(attr, '') if hasattr(self.elem, 'get') else ''
                            def get_text(self, strip=False):
                                # Пробуем найти название книги в элементе результата
                                title_elem = self.item.find('span', class_='Title')
                                if title_elem:
                                    title = title_elem.get_text(strip=strip)
                                    title = re.sub(r'<[^>]+>', '', title)
                                    return title
                                return ''
                            def find_parent(self, *args):
                                return self.item if args else None
                        view_link = DataLinkLink(data_link, data_link_elem, item)
                        print(f"[DEBUG] RMЭБ: найдена ссылка на книгу в data-link: {data_link}")
                        break
            
            if view_link:
                all_links.append(view_link)
            else:
                print(f"[DEBUG] RMЭБ: ссылка /book/ не найдена в элементе {i+1}, ищем data-id")
            
            # Также проверяем элементы с data-id - можем построить URL
            # Ищем в разных местах: в result-access-link, search-items и т.д.
            data_id_elem = None
            book_id = None
            
            # Сначала ищем data-id в любом месте элемента результата
            all_data_id_elems = item.find_all(attrs={'data-id': True})
            if all_data_id_elems:
                data_id_elem = all_data_id_elems[0]
                book_id = data_id_elem.get('data-id', '')
                print(f"[DEBUG] RMЭБ: найден data-id напрямую в элементе {i+1}: {book_id}")
            
            # Если не нашли, ищем в result-access-link
            if not book_id:
                access_link = item.find('div', class_='result-access-link')
                if access_link:
                    # Ищем все элементы с data-id внутри result-access-link
                    access_data_id_elems = access_link.find_all(attrs={'data-id': True})
                    if access_data_id_elems:
                        data_id_elem = access_data_id_elems[0]
                        book_id = data_id_elem.get('data-id', '')
                        print(f"[DEBUG] RMЭБ: найден data-id в result-access-link элемента {i+1}: {book_id}")
            
            # Если все еще не нашли, ищем search-items
            if not book_id:
                search_items = item.find('li', class_='search-items')
                if search_items and search_items.get('data-id'):
                    data_id_elem = search_items
                    book_id = search_items.get('data-id', '')
                    print(f"[DEBUG] RMЭБ: найден data-id в search-items элемента {i+1}: {book_id}")
            
            # Если не нашли data-id, пробуем извлечь из onclick или других мест
            if not book_id:
                # Ищем элементы с onclick, которые могут содержать ID книги
                onclick_elems = item.find_all(attrs={'onclick': True})
                for onclick_elem in onclick_elems:
                    onclick = onclick_elem.get('onclick', '')
                    # Ищем паттерны типа /book/123 или book/123
                    book_match = re.search(r'/book/(\d+)', onclick)
                    if book_match:
                        book_id = book_match.group(1)
                        data_id_elem = onclick_elem
                        print(f"[DEBUG] RMЭБ: найден book_id из onclick элемента {i+1}: {book_id}")
                        break
                
                # Если все еще не нашли, пробуем найти в тексте элемента (может быть скрыт в HTML)
                if not book_id:
                    item_html = str(item)
                    # Ищем паттерн /book/123 в HTML
                    book_match = re.search(r'/book/(\d+)', item_html)
                    if book_match:
                        book_id = book_match.group(1)
                        print(f"[DEBUG] RMЭБ: найден book_id из HTML элемента {i+1}: {book_id}")
            
            # Если нашли book_id, создаем ссылку
            if book_id:
                # Создаем искусственную ссылку
                class DataIdLink:
                    def __init__(self, book_id, elem, item):
                        self.book_id = book_id
                        self.elem = elem
                        self.item = item
                        self.href = f"/book/{book_id}"
                    def get(self, attr):
                        if attr == 'href':
                            return self.href
                        if hasattr(self.elem, 'get'):
                            return self.elem.get(attr, '')
                        return getattr(self.elem, attr, '') if hasattr(self.elem, attr) else ''
                    def get_text(self, strip=False):
                        # Пробуем найти название книги в элементе результата
                        title_elem = self.item.find('span', class_='Title')
                        if title_elem:
                            title = title_elem.get_text(strip=strip)
                            # Убираем HTML теги
                            title = re.sub(r'<[^>]+>', '', title)
                            return title
                        # Пробуем data-title
                        if hasattr(self.elem, 'get'):
                            return self.elem.get('data-title', '')
                        return ''
                    def find_parent(self, *args):
                        return self.item if args else None
                all_links.append(DataIdLink(book_id, data_id_elem or item, item))
                print(f"[DEBUG] RMЭБ: создана ссылка из data-id: /book/{book_id}, добавлена в all_links")
            else:
                print(f"[DEBUG] RMЭБ: data-id не найден в элементе {i+1}")
                # Логируем все элементы с data-* атрибутами для отладки
                all_data_attrs = item.find_all(lambda tag: any(attr.startswith('data-') for attr in tag.attrs.keys()))
                if all_data_attrs:
                    print(f"[DEBUG] RMЭБ: найдены элементы с data-* атрибутами: {len(all_data_attrs)}")
                    for idx, elem in enumerate(all_data_attrs[:3]):  # Первые 3
                        print(f"[DEBUG] RMЭБ: data-элемент {idx+1}: {dict(elem.attrs)}")

        # Ищем все возможные ссылки - включая элементы с onclick, которые могут содержать URL
        for elem in onclick_elements:
            onclick = elem.get('onclick', '')
            # Пробуем извлечь URL из onclick (например, window.location.href = "...")
            url_match = re.search(r'["\']([^"\']+)["\']', onclick)
            if url_match:
                url = url_match.group(1)
                if url.startswith('http') or url.startswith('/'):
                    # Создаем временную ссылку
                    class OnClickLink:
                        def __init__(self, url, elem):
                            self.url = url
                            self.elem = elem
                        def get(self, attr):
                            if attr == 'href':
                                return self.url
                            return self.elem.get(attr, '')
                        def get_text(self, strip=False):
                            return self.elem.get_text(strip=strip)
                        def find_parent(self, *args):
                            return self.elem.find_parent(*args)
                    all_links.append(OnClickLink(url, elem))

        # Фильтруем ссылки - ищем только ссылки на книги
        book_links = []
        
        # Паттерны URL, которые могут быть ссылками на книги
        book_url_patterns = [
            '/book/', '/books/', '/publication/', '/publications/', '/resource/', '/resources/',
            '/view/', '/read/', '/download/', '/item/', '/detail/', '/show/',
            '/id/', '/book_id/', '/publication_id/', '/resource_id/',
        ]
        
        # Исключаем навигационные и служебные ссылки
        excluded_paths = [
            '/', '/search', '/home', '/index', '/about', '/contact', '/partners', '/users',
            '/reports/', '/blogs', '/statistics', '/newfeedback', '/ulogin', '/login', '/register',
            '/language', '/magazines', '/lectures', '/resources', '/publications/authors',
            '/search/resources', '/search?',
        ]
        
        excluded_keywords = [
            'javascript:', 'mailto:', 'tel:', 'sms:', 'callto:',
            'facebook.com', 'twitter.com', 'instagram.com', 'vk.com', 'youtube.com', 'telegram.org',
            'edurk.kz',  # Внешний сайт
        ]
        
        # Логируем все найденные ссылки для анализа
        print(f"[DEBUG] RMЭБ: анализируем {len(all_links)} ссылок...")
        for i, link in enumerate(all_links[:30]):
            href = link.get('href') if hasattr(link, 'get') else getattr(link, 'url', getattr(link, 'href', ''))
            text = link.get_text(strip=True)[:50] if hasattr(link, 'get_text') else ''
            print(f"[DEBUG] RMЭБ: ссылка {i+1}: href={str(href)[:100]}, text={text}")
        
        for link in all_links:
            href = link.get('href') if hasattr(link, 'get') else getattr(link, 'url', getattr(link, 'href', ''))
            if not href or href == '#':
                continue
            
            href_str = str(href).strip()
            href_lower = href_str.lower()
            
            # Пропускаем пустые ссылки и якоря
            if not href_str or href_str.startswith('#'):
                continue
            
            # Пропускаем явно служебные ссылки
            if any(keyword in href_lower for keyword in excluded_keywords):
                continue
            
            # Пропускаем навигационные пути
            if any(path in href_lower for path in excluded_paths):
                continue
            
            # Пропускаем очень короткие ссылки (вероятно, не на книги)
            if len(href_str) < 10 and href_str.startswith('/'):
                continue
            
            # Пропускаем ссылки на главную страницу
            if href_lower in ['https://rmebrk.kz', 'https://rmebrk.kz/', 'http://rmebrk.kz', 'http://rmebrk.kz/']:
                continue
            
            # Ищем ссылки, которые выглядят как ссылки на книги
            is_book_link = False
            
            # Приоритет: ссылки вида /book/{id}
            if re.match(r'^/book/\d+', href_str):
                is_book_link = True
                print(f"[DEBUG] RMЭБ: найдена ссылка на книгу /book/: {href_str[:100]}")
            
            # Проверяем паттерны URL книг
            elif any(pattern in href_lower for pattern in book_url_patterns):
                is_book_link = True
                print(f"[DEBUG] RMЭБ: найдена ссылка на книгу по паттерну: {href_str[:100]}")
            
            # Проверяем, содержит ли URL ID (цифры в пути) и находится в элементе результата
            elif re.search(r'/\d+', href_str) and len(href_str) > 15:
                # Проверяем, находится ли ссылка в элементе результата (list-group-item)
                if hasattr(link, 'find_parent'):
                    result_item = link.find_parent('li', class_=lambda x: x and 'list-group-item' in str(x))
                    if result_item:
                        is_book_link = True
                        print(f"[DEBUG] RMЭБ: найдена ссылка с ID в результате: {href_str[:100]}")
            
            # Проверяем, находится ли ссылка в контейнере результатов (не в навигации)
            if not is_book_link and hasattr(link, 'find_parent'):
                parent = link.find_parent(['nav', 'header', 'footer', 'aside'])
                if not parent:  # Не в навигации
                    # Если ссылка достаточно длинная и не в навигации, возможно это книга
                    if len(href_str) > 20:
                        is_book_link = True
            
            if is_book_link:
                book_links.append(link)
            else:
                print(f"[DEBUG] RMЭБ: ссылка пропущена (не похожа на книгу): {href_str[:100]}")
        
        # Добавляем элементы с data-link как ссылки на книги
        for elem in data_link_elements:
            data_link = elem.get('data-link', '')
            if data_link and not data_link.startswith('#') and not any(kw in data_link.lower() for kw in excluded_keywords):
                # Создаем простой объект-ссылку
                class FakeLink:
                    def __init__(self, href, elem):
                        self.href = href
                        self.elem = elem
                    def get(self, attr):
                        if attr == 'href':
                            return self.href
                        return self.elem.get(attr, '')
                    def get_text(self, strip=False):
                        return self.elem.get_text(strip=strip)
                    def find_parent(self, *args):
                        return self.elem.find_parent(*args)
                book_links.append(FakeLink(data_link, elem))
        
        print(f"[DEBUG] RMЭБ: после фильтрации осталось {len(book_links)} ссылок на книги")

        print(f"[DEBUG] RMЭБ: обрабатываем {len(book_links)} потенциальных ссылок на книги")
        
        for i, link in enumerate(book_links[:max_results * 5]):  # Берем больше для фильтрации
            try:
                # Получаем href разными способами
                href = None
                if hasattr(link, 'get'):
                    href = link.get('href')
                elif hasattr(link, 'href'):
                    href = link.href
                elif hasattr(link, 'url'):
                    href = link.url
                elif hasattr(link, 'get_attribute'):
                    href = link.get_attribute('href')
                
                if not href:
                    print(f"[DEBUG] RMЭБ: ссылка {i+1} - нет href, пропускаем")
                    continue

                href_str = str(href).strip()
                if not href_str or href_str == '#':
                    continue

                # Получаем полный URL книги
                book_url = urljoin(base_url_clean, href_str)
                
                # Пропускаем дубликаты
                if any(r.get('url') == book_url for r in results):
                    print(f"[DEBUG] RMЭБ: ссылка {i+1} - дубликат, пропускаем: {book_url[:80]}")
                    continue

                # Получаем название книги
                title = ''
                
                # Сначала пробуем найти элемент результата (list-group-item)
                result_item = None
                if hasattr(link, 'find_parent'):
                    result_item = link.find_parent('li', class_=lambda x: x and 'list-group-item' in str(x))
                elif hasattr(link, 'parent'):
                    result_item = link.parent.find('li', class_=lambda x: x and 'list-group-item' in str(x)) if hasattr(link.parent, 'find') else None
                elif hasattr(link, 'elem'):
                    result_item = link.elem.find_parent('li', class_=lambda x: x and 'list-group-item' in str(x)) if hasattr(link.elem, 'find_parent') else None
                
                if result_item:
                    # Ищем название в <span class="Title">
                    title_elem = result_item.find('span', class_='Title')
                    if title_elem:
                        title = title_elem.get_text(strip=True)
                        # Убираем HTML теги из названия (например, <b>)
                        title = re.sub(r'<[^>]+>', '', title)
                    else:
                        # Пробуем найти в data-title
                        data_id_elem = result_item.find(attrs={'data-title': True})
                        if data_id_elem:
                            title = data_id_elem.get('data-title', '')
                
                # Если не нашли через result_item, пробуем другие способы
                if not title or len(title) < 3:
                    if hasattr(link, 'get_text'):
                        title = link.get_text(strip=True)
                    elif hasattr(link, 'text'):
                        title = link.text
                    elif hasattr(link, 'elem'):
                        title = link.elem.get_text(strip=True) if hasattr(link.elem, 'get_text') else ''
                    
                    # Пробуем найти data-title
                    if (not title or len(title) < 3) and hasattr(link, 'elem'):
                        title = link.elem.get('data-title', '') if hasattr(link.elem, 'get') else ''
                
                if not title or len(title) < 3:
                    # Пробуем найти заголовок рядом
                    parent = None
                    if hasattr(link, 'find_parent'):
                        parent = link.find_parent(['li', 'div', 'article', 'section', 'tr', 'td', 'span'])
                    elif hasattr(link, 'parent'):
                        parent = link.parent
                    elif hasattr(link, 'elem'):
                        parent = link.elem.find_parent(['li', 'div', 'article', 'section', 'tr', 'td', 'span']) if hasattr(link.elem, 'find_parent') else None
                    
                    if parent:
                        # Ищем span.Title
                        title_elem = parent.find('span', class_='Title')
                        if title_elem:
                            title = title_elem.get_text(strip=True)
                            title = re.sub(r'<[^>]+>', '', title)
                        else:
                            title_elem = parent.find(['h1', 'h2', 'h3', 'h4', 'h5', 'strong', 'span', 'p', 'a', 'div'])
                            if title_elem:
                                title = title_elem.get_text(strip=True)
                            else:
                                # Берем весь текст родителя
                                title = parent.get_text(strip=True)
                
                # Если все еще нет названия, пробуем взять из атрибута title или data-*
                if not title or len(title) < 3:
                    if hasattr(link, 'get'):
                        title = link.get('title') or link.get('data-title') or link.get('data-name') or ''
                    elif hasattr(link, 'elem'):
                        title = link.elem.get('title') or link.elem.get('data-title') or link.elem.get('data-name') or ''
                    title = title.strip()

                if not title or len(title) < 3:
                    # Используем URL как название, если ничего не найдено
                    url_parts = href_str.split('/')
                    if url_parts:
                        last_part = url_parts[-1]
                        title = last_part.replace('-', ' ').replace('_', ' ').replace('.html', '').replace('.php', '').title()
                    if len(title) < 3:
                        title = f"Книга {i+1}"

                print(f"[DEBUG] RMЭБ: элемент {i+1} - название: {title[:80]}, ссылка: {book_url[:100]}")

                results.append({
                    "title": title,
                    "url": book_url,
                    "status": "success",
                    "note": "Республиканская Межвузовская Электронная Библиотека - бесплатный доступ к учебникам",
                    "source": "rmebrk"
                })
                
                if len(results) >= max_results:
                    break

            except Exception as e:
                print(f"[DEBUG] RMЭБ: ошибка обработки ссылки {i+1}: {e}")
                import traceback
                print(f"[DEBUG] RMЭБ: traceback: {traceback.format_exc()}")
                continue

        print(f"[DEBUG] RMЭБ: собрано {len(results)} результатов")
        return results

    except Exception as e:
        print(f"[DEBUG] RMЭБ: ошибка поиска: {e}")
        return []


def search_urait_multiple_results(query: str, max_results: int = 5) -> List[Dict[str, Any]]:
    """Ищет несколько результатов на Юрайт"""
    search_url = f"https://urait.ru/search?words={quote_plus(query)}"
    print(f"[DEBUG] Юрайт поиск множественных результатов: {search_url}")
    results = []
    
    try:
        response = requests.get(search_url, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        
        # Ищем все ссылки на книги
        book_links = soup.select("a[href*='/book/']")
        print(f"[DEBUG] Юрайт найдено ссылок на книги: {len(book_links)}")
        
        seen_urls = set()
        for link in book_links[:max_results * 2]:  # Берем больше чтобы отфильтровать дубли
            href = link.get("href", "")
            if not href or href in seen_urls:
                continue
            seen_urls.add(href)
            
            # Получаем название книги
            title = link.get_text(strip=True)
            if not title:
                # Ищем заголовок рядом
                parent = link.find_parent()
                if parent:
                    title_elem = parent.find(['h1', 'h2', 'h3', 'h4', 'h5'])
                    if title_elem:
                        title = title_elem.get_text(strip=True)
            
            if not title:
                title = f"Учебник по предмету: {query}"
            
            full_url = urljoin("https://urait.ru/", href) if not href.startswith("http") else href
            
            results.append({
                "title": title,
                "url": full_url,
                "status": "success",
                "note": "Электронная библиотека Юрайт - полный доступ к учебнику",
                "source": "urait"
            })
            
            if len(results) >= max_results:
                break
                
        print(f"[DEBUG] Юрайт собрано результатов: {len(results)}")
        return results
        
    except Exception as e:
        print(f"[DEBUG] Юрайт ошибка множественного поиска: {e}")
        return []


def search_urait_viewer_link(query: str) -> Optional[str]:
    """Оставляем для обратной совместимости"""
    results = search_urait_multiple_results(query, 1)
    return results[0]["url"] if results else None


def get_urait_book_title(book_url: str, fallback_subject: str) -> str:
    """Получает название книги с страницы Юрайт"""
    try:
        response = requests.get(book_url, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        
        # Ищем заголовок книги
        title_element = soup.select_one("h1.book-title, .book-header__title, h1")
        if title_element:
            title = title_element.get_text().strip()
            if title and len(title) > 5:  # Проверяем что заголовок не пустой
                return title
        
        # Если не нашли заголовок, ищем в meta тегах
        meta_title = soup.select_one("meta[property='og:title']")
        if meta_title and meta_title.get("content"):
            return meta_title["content"].strip()
            
    except Exception:
        pass
    
    # Если ничего не нашли, возвращаем предмет с пометкой
    return f"Учебник по предмету: {fallback_subject}"


def iprbookshop_search_url(query: str) -> str:
    return "https://www.iprbookshop.ru/586.html?title=" + quote_plus(query)


def fetch_iprbookshop_reader(subject: str) -> Optional[Dict[str, Any]]:
    """Поиск на IPRbooks через requests для совместимости с Vercel"""
    print(f"[DEBUG] IPRbooks поиск через requests: '{subject}'")
    
    # Таймауты оптимизированы для Vercel (serverless функции имеют ограничение ~10 сек)
    VERCEL_TIMEOUT = 8  # Уменьшаем таймаут для Vercel
    
    try:
        # Используем requests с более реалистичными заголовками для Vercel
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'ru-RU,ru;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Cache-Control': 'max-age=0'
        })

        # Загружаем предоставленные cookies для авторизованного доступа
        for domain, cookies in IPRBOOKSHOP_COOKIES.items():
            for name, value in cookies.items():
                session.cookies.set(name, value, domain=domain, path="/")
        
        base_url = "https://www.iprbookshop.ru/"
        search_url = urljoin(base_url, "586.html")

        # СНАЧАЛА пробуем AJAX API (быстрее и надежнее для Vercel)
        print("[DEBUG] IPRbooks: пробуем AJAX API сначала")
        ajax_results = fetch_iprbookshop_ajax_results(session, subject, base_url)
        if ajax_results:
            print(f"[DEBUG] IPRbooks: AJAX API вернул {len(ajax_results)} результатов")
            book_elements = ajax_results
        else:
            print("[DEBUG] IPRbooks: AJAX API не вернул результатов, пробуем HTML поиск")
            
            # Fallback: получаем главную страницу для установки сессии/куки
            try:
                main_response = session.get(base_url, timeout=VERCEL_TIMEOUT)
                print(f"[DEBUG] IPRbooks: главная страница - статус {main_response.status_code}")
            except Exception as e:
                print(f"[DEBUG] IPRbooks: ошибка получения главной страницы: {e}")

            # Основной запрос: POST с данными поиска
            search_data = {'pagetitle': subject, 'submit': 'Применить'}
            headers = {
                'Referer': base_url,
                'Origin': base_url.rstrip('/')
            }
            print(f"[DEBUG] IPRbooks: отправляем POST запрос с данными: {search_data}")
            
            try:
                response = session.post(search_url, data=search_data, headers=headers, timeout=VERCEL_TIMEOUT, allow_redirects=True)
                print(f"[DEBUG] IPRbooks: POST ответ - статус {response.status_code}, URL: {response.url}")
            except requests.exceptions.Timeout:
                print("[DEBUG] IPRbooks: таймаут POST запроса")
                return None
            except requests.exceptions.RequestException as e:
                print(f"[DEBUG] IPRbooks: ошибка POST запроса: {e}")
                return None

            if response.status_code != 200:
                print(f"[DEBUG] IPRbooks: неожиданный статус {response.status_code}")
                return None

            preview_text = response.text[:800].replace('\n', ' ').strip()
            print(f"[DEBUG] IPRbooks: превью HTML: {preview_text}")
            if not preview_text:
                print("[DEBUG] IPRbooks: тело ответа пустое")
                return None

            if "auth" in response.url.lower() or "login" in response.url.lower():
                print("[DEBUG] IPRbooks: перенаправление на авторизацию")
                return None
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Отладочная информация о содержимом страницы
            page_title = soup.find('title')
            print(f"[DEBUG] IPRbooks: заголовок страницы: {page_title.get_text() if page_title else 'Не найден'}")
            
            # Проверяем наличие контейнера результатов (может быть пустым, если загружается через AJAX)
            results_container = soup.find('div', {'id': 'ajaxContentBooks'})
            if results_container:
                container_html_preview = results_container.prettify()[:500].replace('\n', ' ').strip()
                print(f"[DEBUG] IPRbooks: контейнер ajaxContentBooks найден, содержимое: {container_html_preview}")
            
            # Ищем результаты поиска - пробуем разные селекторы
            book_elements = soup.select('div.row.row-book')
            print(f"[DEBUG] IPRbooks: div.row.row-book найдено: {len(book_elements)}")

            # Если не нашли стандартные элементы, пробуем другие селекторы
            if not book_elements:
                book_elements = soup.select('.book-item, .search-result, .publication, .row-book')
                print(f"[DEBUG] IPRbooks: альтернативные селекторы найдено: {len(book_elements)}")
            
            # Пробуем найти в контейнере AJAX
            if not book_elements and results_container:
                book_elements = results_container.select('div.row.row-book, .book-item, .row-book')
                print(f"[DEBUG] IPRbooks: в контейнере ajaxContentBooks найдено: {len(book_elements)}")
            
            # Пробуем найти любые ссылки на книги
            if not book_elements:
                book_links = soup.select('a[href*="/book/"], a[href*="/publication/"]')
                print(f"[DEBUG] IPRbooks: найдено ссылок на книги: {len(book_links)}")
                if book_links:
                    # Создаем фиктивные элементы из ссылок
                    for link in book_links[:10]:  # Берем первые 10
                        parent = link.find_parent(['div', 'li', 'article'])
                        if parent:
                            book_elements.append(parent)
                    print(f"[DEBUG] IPRbooks: создано элементов из ссылок: {len(book_elements)}")

            if not book_elements:
                # Логируем структуру страницы для отладки
                page_structure = soup.find_all(['div', 'section', 'article'], class_=True)[:10]
                print(f"[DEBUG] IPRbooks: структура страницы (первые 10 элементов с классами):")
                for elem in page_structure:
                    print(f"  - {elem.name}.{'.'.join(elem.get('class', []))}")
                print("[DEBUG] IPRbooks: результаты не найдены в HTML")
                return None
            
            # Проверяем релевантность HTML результатов
            if book_elements:
                subject_words = [w.lower() for w in subject.split() if len(w) > 2]
                if subject_words:
                    relevant_count = 0
                    for book_elem in book_elements[:5]:
                        try:
                            title_link = book_elem.select_one('h4 a')
                            if title_link:
                                title_text = title_link.get_text(strip=True).lower()
                                matches = sum(1 for word in subject_words if word in title_text)
                                if matches > 0:
                                    relevant_count += 1
                        except Exception:
                            pass
                    
                    if relevant_count == 0 and len(book_elements) > 0:
                        print(f"[DEBUG] IPRbooks HTML: предупреждение - результаты не релевантны запросу '{subject}' (0/{min(5, len(book_elements))} релевантных)")
                        # Но все равно используем их, так как это fallback
        
        # Применяем скоринг для выбора лучшей книги
        best_book = None
        best_score = -100
        best_element = None
        
        for book_elem in book_elements:
            try:
                title_link = book_elem.select_one('h4 a')
                if not title_link:
                    continue
                
                title_text = title_link.get_text(strip=True)
                title_lower = title_text.lower()
                subject_lower = subject.lower()
                
                # Scoring system for book relevance
                score = 0
                
                # Exact match gets highest score
                if subject_lower in title_lower:
                    score += 10
                
                # Проверяем совпадение ключевых слов из запроса
                subject_words = [w for w in subject_lower.split() if len(w) > 2]
                title_words = title_lower.split()
                
                # Подсчитываем совпадения слов
                matched_words = sum(1 for word in subject_words if word in title_lower)
                if matched_words > 0:
                    score += matched_words * 3  # +3 за каждое совпавшее слово
                
                # Бонус за совпадение важных слов (первые слова запроса обычно важнее)
                if subject_words:
                    first_word = subject_words[0]
                    if first_word in title_lower:
                        score += 5
                
                # Проверяем совпадение по частям (для составных терминов)
                # Например, "анатомия человека" должно совпадать с "анатомия" и "человек"
                important_parts = [w for w in subject_words if len(w) > 4]  # Длинные слова важнее
                for part in important_parts:
                    if part in title_lower:
                        score += 4
                
                # Penalty for completely unrelated topics
                unrelated_words = ["программирование", "информатика", "математика", "физика", "химия", "английский", "язык"]
                if any(word in title_lower for word in unrelated_words) and not any(word in subject_lower for word in unrelated_words):
                    score -= 15
                
                # Strong penalty for automation books when not searching for automation
                automation_words = ["автоматизация", "автоматизированный", "automation", "автомат", "машиностроение"]
                automation_in_title = any(word in title_lower for word in automation_words)
                automation_in_subject = any(word in subject_lower for word in ["автомат", "машин", "производств"])
                
                if automation_in_title and not automation_in_subject:
                    score -= 25  # Very strong penalty
                
                # Penalty for books that start with "А" when searching for something else (alphabetical sorting issue)
                if title_lower.startswith('а') and not subject_lower.startswith('а'):
                    # Если запрос не начинается с "А", но книга начинается - это может быть алфавитная сортировка
                    # Небольшой штраф, но не критичный
                    if score < 5:  # Только если уже низкий балл
                        score -= 2
                
                print(f"[DEBUG] IPRbooks: книга '{title_text[:50]}...' - балл: {score}")
                
                if score > best_score:
                    best_score = score
                    best_book = title_text
                    best_element = book_elem
                    
            except Exception as e:
                print(f"[DEBUG] IPRbooks ошибка обработки элемента: {e}")
                continue
        
        print(f"[DEBUG] IPRbooks: лучшая книга выбрана с баллом {best_score}: '{best_book[:80] if best_book else 'None'}...'")
        
        # Если не нашли хорошую книгу (балл слишком низкий), пробуем следующий вариант поиска
        if best_element is None or best_score < 3:
            print(f"[DEBUG] IPRbooks: не найдено подходящих книг (лучший балл: {best_score}), берем первую")
            if book_elements:
                best_element = book_elements[0]
                title_link = best_element.select_one('h4 a')
                if title_link:
                    best_book = title_link.get_text(strip=True)
                else:
                    # Пробуем получить из data-атрибутов
                    book_id = best_element.get('data-book-id')
                    pagetitle = best_element.get('data-pagetitle')
                    if pagetitle:
                        best_book = pagetitle
        
        if not best_element:
            return None
        
        # Получаем ссылку на книгу
        title_link = best_element.select_one('h4 a')
        if not title_link:
            # Если ссылка не найдена, пробуем получить из data-атрибутов
            book_id = best_element.get('data-book-id')
            if book_id:
                detail_url = f"https://www.iprbookshop.ru/{book_id}.html"
            else:
                return None
        else:
            detail_href = title_link.get('href')
            if not detail_href or detail_href.startswith('javascript'):
                # Пробуем получить из data-атрибутов
                book_id = best_element.get('data-book-id')
                if book_id:
                    detail_url = f"https://www.iprbookshop.ru/{book_id}.html"
                else:
                    return None
            else:
                # Если ссылка относительная, делаем её абсолютной
                if detail_href.startswith('/'):
                    detail_url = urljoin("https://www.iprbookshop.ru/", detail_href)
                elif detail_href.startswith('http'):
                    detail_url = detail_href
                else:
                    detail_url = urljoin("https://www.iprbookshop.ru/", detail_href)
        
        # Пытаемся получить прямую ссылку для чтения
        reader_url = detail_url
        note_parts = []
        
        try:
            print(f"[DEBUG] IPRbooks: получаем детали книги: {detail_url}")
            detail_headers = {
                'Referer': search_url,
                'Origin': base_url
            }
            detail_response = session.get(detail_url, headers=detail_headers, timeout=VERCEL_TIMEOUT)
            
            if detail_response.status_code == 200:
                detail_preview = detail_response.text[:600].replace('\n', ' ').strip()
                print(f"[DEBUG] IPRbooks: превью страницы книги: {detail_preview}")
                detail_soup = BeautifulSoup(detail_response.text, 'html.parser')
                
                # Ищем кнопку "Читать" с разными селекторами
                read_button = (
                    detail_soup.select_one('a.btn-read') or
                    detail_soup.select_one('a[href*="epd-reader"]') or
                    detail_soup.select_one('a[href*="reader"]') or
                    detail_soup.select_one('.read-btn')
                )
                
                if read_button and read_button.get('href'):
                    reader_href = read_button.get('href')
                    reader_url = urljoin(detail_url, reader_href)
                    print(f"[DEBUG] IPRbooks: найдена прямая ссылка для чтения: {reader_url}")
                else:
                    print("[DEBUG] IPRbooks: кнопка 'Читать' не найдена, используем ссылку на страницу")
                
                # Получаем информацию о публикации
                pub_data_elements = detail_soup.select('div.pub-data, .publication-info, .book-info')
                for elem in pub_data_elements:
                    text = elem.get_text(strip=True)
                    if text and len(text) > 5:
                        note_parts.append(text)
                
                if not note_parts:
                    year_elem = detail_soup.select_one('.year, .publication-year')
                    if year_elem:
                        note_parts.append(f"Год: {year_elem.get_text(strip=True)}")
                    pages_elem = detail_soup.select_one('.pages, .page-count')
                    if pages_elem:
                        note_parts.append(f"Страниц: {pages_elem.get_text(strip=True)}")
            else:
                print(f"[DEBUG] IPRbooks: ошибка получения деталей - статус {detail_response.status_code}")
        
        except Exception as e:
            print(f"[DEBUG] IPRbooks: не удалось получить детали книги: {e}")
        
        if not note_parts:
            note_parts = ["IPRbooks - Электронная библиотека"]
        
        return {
            "url": reader_url,
            "title": best_book or f"Учебник по предмету: {subject}",
            "status": "success",
            "note": "; ".join(note_parts) if note_parts else "Читать на IPRbooks",
            "source": "iprbookshop",
            "multiple": False
        }
        
    except requests.exceptions.Timeout:
        print(f"[DEBUG] IPRbooks: таймаут соединения")
        return None
    except requests.exceptions.ConnectionError:
        print(f"[DEBUG] IPRbooks: ошибка соединения")
        return None
    except requests.exceptions.RequestException as e:
        print(f"[DEBUG] IPRbooks: ошибка запроса: {e}")
        return None
    except Exception as e:
        print(f"[DEBUG] IPRbooks: неожиданная ошибка: {e}")
        return None


# Функция get_first_iprbookshop_result больше не нужна, так как основная функция теперь использует requests


def fetch_iprbookshop_ajax_results(session: requests.Session, subject: str, base_url: str) -> List[Tag]:
    """Повторяет AJAX-запрос /107257, который делает фронтенд IPR SMART"""
    VERCEL_TIMEOUT = 7  # Уменьшенный таймаут для Vercel
    
    ajax_url = urljoin(base_url, "107257")
    # Пробуем добавить поисковый запрос в URL параметры тоже
    params = {"page": 1}
    
    # ВАЖНО: Сначала загружаем страницу поиска, чтобы получить правильную сессию
    search_page_url = urljoin(base_url, "586.html")
    input_name = 'pagetitle'  # По умолчанию
    
    try:
        print(f"[DEBUG] IPRbooks AJAX: загружаем страницу поиска для установки сессии")
        search_page_response = session.get(search_page_url, timeout=VERCEL_TIMEOUT)
        if search_page_response.status_code != 200:
            print(f"[DEBUG] IPRbooks AJAX: ошибка загрузки страницы поиска: {search_page_response.status_code}")
        else:
            # Парсим страницу, чтобы найти форму поиска и её параметры
            search_soup = BeautifulSoup(search_page_response.text, 'html.parser')
            # Ищем поле поиска - может быть pagetitle, title, или другое имя
            search_input = (
                search_soup.find('input', {'id': 'pagetitle'}) or 
                search_soup.find('input', {'name': 'pagetitle'}) or
                search_soup.find('input', {'id': 'title'}) or
                search_soup.find('input', {'name': 'title'})
            )
            if search_input:
                input_name = search_input.get('name') or search_input.get('id') or 'pagetitle'
                print(f"[DEBUG] IPRbooks AJAX: найдено поле поиска с именем: {input_name}")
            
            # ВАЖНО: Отправляем POST запрос на страницу поиска, чтобы установить поисковый запрос в сессии
            # Это может быть необходимо для того, чтобы AJAX запрос работал правильно
            try:
                print(f"[DEBUG] IPRbooks AJAX: отправляем POST на страницу поиска для установки запроса '{subject}'")
                # Используем правильное имя поля из формы
                form_data = {
                    input_name: subject.strip(),
                    'submit': 'Применить'
                }
                form_response = session.post(search_page_url, data=form_data, timeout=VERCEL_TIMEOUT, allow_redirects=True)
                print(f"[DEBUG] IPRbooks AJAX: POST на страницу поиска - статус {form_response.status_code}")
                
                # Обновляем cookies из ответа, чтобы сохранить сессию
                if form_response.cookies:
                    session.cookies.update(form_response.cookies)
                    print(f"[DEBUG] IPRbooks AJAX: обновлены cookies из ответа формы")
                
                # Проверяем, что поиск действительно установлен - ищем наш запрос в ответе
                if subject.lower() not in form_response.text.lower()[:5000]:
                    print(f"[DEBUG] IPRbooks AJAX: предупреждение - запрос '{subject}' не найден в ответе формы поиска")
                
                # Небольшая задержка для обработки на сервере
                import time
                time.sleep(0.5)
            except Exception as e:
                print(f"[DEBUG] IPRbooks AJAX: ошибка отправки формы поиска: {e}")
    except Exception as e:
        print(f"[DEBUG] IPRbooks AJAX: ошибка загрузки страницы поиска: {e}")
    
    # Формируем payload - используем оба варианта имени поля для надежности
    # AJAX API может использовать pagetitle, но форма использует title
    base_payload = {
        "action": "getPublications",
        "pagetitle": subject.strip(),  # AJAX API обычно использует pagetitle
    }
    # Если форма использует 'title', добавляем его тоже
    if input_name == 'title':
        base_payload["title"] = subject.strip()
        print(f"[DEBUG] IPRbooks AJAX: добавляем поле 'title' в payload (форма использует это имя)")
    headers = {
        "Referer": search_page_url,
        "Origin": base_url.rstrip("/"),
        "X-Requested-With": "XMLHttpRequest",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
    }

    # Пробуем варианты поиска в порядке приоритета
    search_variants = [
        {"search_type": 1, "available": 1},  # По названию, доступные
        {"search_type": 2, "available": 1},  # По содержанию, доступные
        {"search_type": 1, "available": 2},  # По названию, все
        {"search_type": 2, "available": 2},  # По содержанию, все
    ]

    try:
        
        for variant in search_variants:
            payload = {**base_payload, **variant}
            print(f"[DEBUG] IPRbooks AJAX: попытка поиска '{subject}' с параметрами {payload}")
            print(f"[DEBUG] IPRbooks AJAX: отправляем POST на {ajax_url} с данными: pagetitle='{payload.get('pagetitle')}', search_type={payload.get('search_type')}, available={payload.get('available')}")
            
            try:
                # Отправляем POST запрос с данными поиска
                response = session.post(
                    ajax_url, 
                    params=params, 
                    data=payload, 
                    headers=headers, 
                    timeout=VERCEL_TIMEOUT
                )
                print(f"[DEBUG] IPRbooks AJAX: статус {response.status_code}, url {response.url}")
                print(f"[DEBUG] IPRbooks AJAX: размер ответа: {len(response.text)} байт")
                
                # Проверяем, что запрос действительно содержит наш поисковый запрос
                response_preview = response.text.lower()[:2000]
                if subject.lower() not in response_preview:
                    # Если в ответе нет нашего запроса, возможно поиск не сработал
                    print(f"[DEBUG] IPRbooks AJAX: предупреждение - в ответе не найден запрос '{subject}'")
                    print(f"[DEBUG] IPRbooks AJAX: превью ответа: {response.text[:500]}")
            except requests.exceptions.Timeout:
                print(f"[DEBUG] IPRbooks AJAX: таймаут для варианта {variant}")
                continue
            except requests.exceptions.RequestException as e:
                print(f"[DEBUG] IPRbooks AJAX: ошибка запроса для варианта {variant}: {e}")
                continue

            if response.status_code != 200:
                print(f"[DEBUG] IPRbooks AJAX: неожиданный статус {response.status_code}")
                continue

            # Парсим JSON ответ
            try:
                data = response.json()
            except ValueError as e:
                print(f"[DEBUG] IPRbooks AJAX: ошибка парсинга JSON: {e}")
                print(f"[DEBUG] IPRbooks AJAX: ответ (первые 1000 символов): {response.text[:1000]}")
                continue

            if not data:
                print("[DEBUG] IPRbooks AJAX: пустой JSON")
                continue

            # Детальное логирование структуры ответа
            print(f"[DEBUG] IPRbooks AJAX: тип ответа: {type(data)}")
            if isinstance(data, dict):
                print(f"[DEBUG] IPRbooks AJAX: ключи в ответе: {list(data.keys())}")
                print(f"[DEBUG] IPRbooks AJAX: success = {data.get('success')}")
                
                # Проверяем данные на релевантность запросу
                data_items = data.get('data', [])
                if data_items and len(data_items) > 0:
                    first_item = data_items[0] if isinstance(data_items[0], dict) else None
                    if first_item:
                        first_title = first_item.get('pagetitle', '')
                        print(f"[DEBUG] IPRbooks AJAX: первый результат: '{first_title[:80]}...'")
                        # Проверяем, содержит ли первый результат наш запрос
                        subject_words = [w.lower() for w in subject.split() if len(w) > 2]
                        title_lower = first_title.lower()
                        matches = sum(1 for word in subject_words if word in title_lower)
                        print(f"[DEBUG] IPRbooks AJAX: совпадений ключевых слов в первом результате: {matches}/{len(subject_words)}")
                        if matches == 0 and len(subject_words) > 0:
                            print(f"[DEBUG] IPRbooks AJAX: ВНИМАНИЕ: результаты не соответствуют запросу '{subject}'!")
                            print(f"[DEBUG] IPRbooks AJAX: возможно, API игнорирует параметр поиска или возвращает общие результаты")
                
                # Логируем первые несколько ключей с их типами
                for key in list(data.keys())[:3]:
                    value = data[key]
                    value_type = type(value).__name__
                    if isinstance(value, (list, tuple)):
                        print(f"[DEBUG] IPRbooks AJAX:   {key} ({value_type}): длина={len(value)}")
                    elif isinstance(value, str):
                        preview = value[:100] if len(value) > 100 else value
                        print(f"[DEBUG] IPRbooks AJAX:   {key} ({value_type}): {preview}")
            elif isinstance(data, list):
                print(f"[DEBUG] IPRbooks AJAX: ответ - список длиной {len(data)}")
                if len(data) > 0:
                    print(f"[DEBUG] IPRbooks AJAX: первый элемент типа: {type(data[0])}")
            else:
                print(f"[DEBUG] IPRbooks AJAX: ответ (первые 500 символов): {str(data)[:500]}")
            
            if isinstance(data, dict) and data.get("success") is False:
                message = data.get('message', 'Неизвестная ошибка')
                print(f"[DEBUG] IPRbooks AJAX: отказ сервера — {message}")
                continue

            # Извлекаем HTML элементы из ответа
            book_elements: List[Tag] = []
            
            # СНАЧАЛА пробуем использовать text_data (HTML контейнер с результатами)
            text_data_found = False
            if isinstance(data, dict) and data.get("text_data"):
                text_data = data.get("text_data")
                if isinstance(text_data, str) and len(text_data) > 100:
                    print(f"[DEBUG] IPRbooks AJAX: проверяем text_data (длина: {len(text_data)})")
                    try:
                        text_soup = BeautifulSoup(text_data, "html.parser")
                        # Ищем элементы книг в text_data
                        found_elements = text_soup.select('div.row.row-book, .row.row-book, div[class*="row-book"], .book-item')
                        if found_elements and len(found_elements) > 0:
                            book_elements.extend(found_elements)
                            text_data_found = True
                            print(f"[DEBUG] IPRbooks AJAX: найдено {len(found_elements)} элементов в text_data")
                        else:
                            print(f"[DEBUG] IPRbooks AJAX: text_data не содержит элементов книг, используем data")
                    except Exception as e:
                        print(f"[DEBUG] IPRbooks AJAX: ошибка парсинга text_data: {e}")
            
            # Если не нашли в text_data, пробуем парсить data (список словарей)
            if not text_data_found:
                data_items = None
                if isinstance(data, dict):
                    data_items = data.get("data")
                elif isinstance(data, list):
                    data_items = data
                
                if not data_items:
                    print(f"[DEBUG] IPRbooks AJAX: нет данных в ответе")
                    continue
                
                print(f"[DEBUG] IPRbooks AJAX: найдено элементов данных: {len(data_items) if isinstance(data_items, (list, tuple)) else 1}")
                
                # Обрабатываем data_items - создаем элементы из словарей
                if not isinstance(data_items, (list, tuple)):
                    data_items = [data_items]
                
                for idx, item in enumerate(data_items):
                    if not item or not isinstance(item, dict):
                        continue
                    
                    # Извлекаем данные из словаря
                    book_id = item.get("id")
                    pagetitle = item.get("pagetitle", "")
                    
                    if not book_id or not pagetitle:
                        continue
                    
                    # Создаем HTML элемент из данных
                    # Формат ссылки на IPRbooks: https://www.iprbookshop.ru/{id}.html
                    book_url = f"https://www.iprbookshop.ru/{book_id}.html"
                    book_html = f'''
                    <div class="row row-book">
                        <h4><a href="{book_url}">{pagetitle}</a></h4>
                    </div>
                    '''
                    
                    try:
                        book_soup = BeautifulSoup(book_html, "html.parser")
                        div = book_soup.select_one('div.row.row-book')
                        if div:
                            # Сохраняем оригинальные данные в атрибутах для дальнейшего использования
                            div['data-book-id'] = str(book_id)
                            div['data-pagetitle'] = pagetitle
                            book_elements.append(div)
                            print(f"[DEBUG] IPRbooks AJAX: создан элемент {idx}: {pagetitle[:50]}...")
                    except Exception as e:
                        print(f"[DEBUG] IPRbooks AJAX: ошибка создания элемента {idx}: {e}")
                        continue

            if book_elements:
                # Проверяем релевантность результатов перед возвратом
                # Если результаты не соответствуют запросу, пробуем следующий вариант
                subject_words = [w.lower() for w in subject.split() if len(w) > 2]
                if subject_words:
                    # Проверяем первые несколько результатов на релевантность
                    relevant_count = 0
                    for book_elem in book_elements[:5]:  # Проверяем первые 5
                        try:
                            title_link = book_elem.select_one('h4 a')
                            if title_link:
                                title_text = title_link.get_text(strip=True).lower()
                                # Проверяем совпадение ключевых слов
                                matches = sum(1 for word in subject_words if word in title_text)
                                if matches > 0:
                                    relevant_count += 1
                        except Exception:
                            pass
                    
                    # Если менее 20% результатов релевантны, считаем что поиск не сработал
                    if relevant_count == 0 and len(book_elements) > 0:
                        print(f"[DEBUG] IPRbooks AJAX: результаты не релевантны запросу '{subject}' (0/{min(5, len(book_elements))} релевантных), пробуем следующий вариант")
                        continue
                    elif relevant_count > 0:
                        print(f"[DEBUG] IPRbooks AJAX: найдено {relevant_count} релевантных результатов из {min(5, len(book_elements))}, используем их")
                
                print(f"[DEBUG] IPRbooks AJAX: итого получено элементов {len(book_elements)}")
                return book_elements

        print("[DEBUG] IPRbooks AJAX: не удалось получить релевантные данные через API ни для одного варианта")
        return []

    except requests.exceptions.RequestException as exc:
        print(f"[DEBUG] IPRbooks AJAX: критическая ошибка запроса {exc}")
        return []
    except ValueError as exc:
        print(f"[DEBUG] IPRbooks AJAX: критическая ошибка парсинга JSON {exc}")
        return []
    except Exception as exc:
        print(f"[DEBUG] IPRbooks AJAX: неожиданная ошибка {exc}")
        import traceback
        print(f"[DEBUG] IPRbooks AJAX: traceback: {traceback.format_exc()}")
        return []


def get_multiple_iprbookshop_results(page, subject: str, results_count: int, max_results: int = 10) -> List[Dict[str, Any]]:
    """Получает несколько результатов с IPRbooks для API следующего варианта"""
    print(f"[DEBUG] IPRbooks: собираем {max_results} результатов из {results_count}")
    
    base_url = "https://www.iprbookshop.ru/"
    books = page.locator("div.row.row-book")
    results = []
    
    for i in range(min(results_count, max_results)):
        try:
            book = books.nth(i)
            title_element = book.locator("h4 a")
            if not title_element.count():
                continue
                
            title_text = title_element.inner_text().strip()
            detail_href = title_element.get_attribute("href")
            
            if not detail_href or detail_href.startswith("javascript"):
                continue
                
            detail_url = urljoin(base_url, detail_href)
            
            results.append({
                "title": title_text,
                "url": detail_url,
                "status": "warning", 
                "note": f"IPRbooks - вариант {i+1}. Требуется проверка доступности",
                "source": "iprbookshop"
            })
            
        except Exception as e:
            print(f"[DEBUG] IPRbooks ошибка получения результата {i}: {e}")
            continue
    
    return results


def search_iprbookshop_multiple_results(subject: str, max_results: int = 10) -> List[Dict[str, Any]]:
    """Ищет несколько результатов на IPRbooks через requests"""
    print(f"[DEBUG] IPRbooks множественный поиск через requests: '{subject}'")
    VERCEL_TIMEOUT = 7  # Уменьшенный таймаут для Vercel
    
    try:
        # Используем requests с теми же заголовками что и в основной функции
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'ru-RU,ru;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Cache-Control': 'max-age=0'
        })
        
        # Загружаем cookies
        for domain, cookies in IPRBOOKSHOP_COOKIES.items():
            for name, value in cookies.items():
                session.cookies.set(name, value, domain=domain, path="/")
        
        base_url = "https://www.iprbookshop.ru/"
        
        # СНАЧАЛА пробуем AJAX API (быстрее и надежнее)
        print("[DEBUG] IPRbooks множественный: пробуем AJAX API")
        ajax_results = fetch_iprbookshop_ajax_results(session, subject, base_url)
        
        book_elements = []
        if ajax_results:
            print(f"[DEBUG] IPRbooks множественный: AJAX API вернул {len(ajax_results)} результатов")
            book_elements = ajax_results
        else:
            # Fallback: HTML поиск
            print("[DEBUG] IPRbooks множественный: AJAX не сработал, пробуем HTML поиск")
            try:
                main_response = session.get(base_url, timeout=VERCEL_TIMEOUT)
                print(f"[DEBUG] IPRbooks множественный: главная страница - статус {main_response.status_code}")
            except Exception as e:
                print(f"[DEBUG] IPRbooks множественный: ошибка получения главной страницы: {e}")
            
            search_url = urljoin(base_url, "586.html")
            search_data = {
                'pagetitle': subject,
                'submit': 'Применить'
            }
            
            try:
                response = session.post(search_url, data=search_data, timeout=VERCEL_TIMEOUT, allow_redirects=True)
                print(f"[DEBUG] IPRbooks множественный: POST ответ - статус {response.status_code}")
            except requests.exceptions.Timeout:
                print("[DEBUG] IPRbooks множественный: таймаут POST запроса")
                return []
            except requests.exceptions.RequestException as e:
                print(f"[DEBUG] IPRbooks множественный: ошибка POST запроса: {e}")
                return []
            
            if response.status_code != 200:
                print(f"[DEBUG] IPRbooks множественный: неожиданный статус {response.status_code}")
                return []
            
            soup = BeautifulSoup(response.text, 'html.parser')
            book_elements = soup.select('div.row.row-book')
            print(f"[DEBUG] IPRbooks найдено элементов книг: {len(book_elements)}")
        
        if not book_elements:
            print("[DEBUG] IPRbooks множественный: результаты не найдены")
            return []
        
        results = []
        
        for i, book_elem in enumerate(book_elements[:max_results]):
            try:
                # Ищем ссылку на книгу
                title_link = book_elem.select_one('h4 a')
                if not title_link:
                    continue
                
                title_text = title_link.get_text(strip=True)
                detail_href = title_link.get('href')
                
                if not detail_href or detail_href.startswith('javascript'):
                    continue
                
                detail_url = urljoin(base_url, detail_href)
                
                # Пытаемся получить прямую ссылку для чтения (с коротким таймаутом)
                reader_url = detail_url
                try:
                    detail_response = session.get(detail_url, timeout=5)
                    if detail_response.status_code == 200:
                        detail_soup = BeautifulSoup(detail_response.text, 'html.parser')
                        
                        # Ищем кнопку "Читать"
                        read_button = (
                            detail_soup.select_one('a.btn-read') or
                            detail_soup.select_one('a[href*="epd-reader"]') or
                            detail_soup.select_one('a[href*="reader"]')
                        )
                        if read_button and read_button.get('href'):
                            reader_href = read_button.get('href')
                            reader_url = urljoin(detail_url, reader_href)
                            print(f"[DEBUG] IPRbooks: найдена прямая ссылка для чтения: {reader_url}")
                
                except Exception as e:
                    print(f"[DEBUG] IPRbooks: не удалось получить прямую ссылку для {i}: {e}")
                
                results.append({
                    "title": title_text,
                    "url": reader_url,
                    "status": "warning", 
                    "note": f"IPRbooks - вариант {i+1}. Требуется проверка доступности",
                    "source": "iprbookshop"
                })
                
            except Exception as e:
                print(f"[DEBUG] IPRbooks ошибка получения результата {i}: {e}")
                continue
        
        print(f"[DEBUG] IPRbooks собрано результатов: {len(results)}")
        return results
        
    except Exception as e:
        print(f"[DEBUG] IPRbooks ошибка множественного поиска: {e}")
        import traceback
        print(f"[DEBUG] IPRbooks traceback: {traceback.format_exc()}")
        return []


def fetch_links_for_subject(subject: str) -> Tuple[str, Dict[str, Any]]:
    info: Dict[str, Any] = {
        "links": [],
        "primary_link": None,
        "status": "warning",
        "note": "",
        "resources": [],
    }

    norm = normalize_subject(subject)
    for rule in KNOWN_RESOURCE_RULES:
        if all(token in norm for token in rule.get("contains", [])):
            resources: List[Dict[str, Any]] = []
            for res in rule.get("resources", []):
                if not res.get("url"):
                    continue
                resources.append(
                    {
                        "title": res.get("title") or res.get("url"),
                        "url": res.get("url"),
                        "status": res.get("status", "success"),
                        "note": res.get("note", ""),
                    }
                )
            if resources:
                info["resources"] = resources
                info["links"] = [res["url"] for res in resources]
                info["primary_link"] = resources[0]["url"]
                info["status"] = resources[0].get("status", "success")
                info["note"] = resources[0].get("note", "")
                if info["status"] == "warning" and not info["note"]:
                    info["note"] = "Проверьте предложенный ресурс"
            return subject, info

    try:
        print(f"[DEBUG] Ищем на RMЭБ: '{subject}'")
        # На Vercel берем меньше результатов
        max_rmebrk_results = 2 if os.getenv('VERCEL') == '1' else 3
        rmebrk_results = search_rmebrk_results(subject, max_rmebrk_results)
        if rmebrk_results:
            print(f"[DEBUG] RMЭБ найдено {len(rmebrk_results)} результатов")
            for result in rmebrk_results[:1]:  # Добавляем только первый результат в links, но все в resources
                info["links"].append(result["url"])
                info["primary_link"] = info["primary_link"] or result["url"]
                info["status"] = "success"
                info["note"] = "Найдено на RMЭБ"
            info["resources"].extend(rmebrk_results)
        else:
            print(f"[DEBUG] RMЭБ не найден для: '{subject}'")
    except Exception as exc:
        print(f"[DEBUG] Ошибка поиска RMЭБ: {exc}")
        # Не падаем с ошибкой, продолжаем поиск на других сайтах

    try:
        print(f"[DEBUG] Ищем на Юрайт: '{subject}'")
        urait_results = search_urait_multiple_results(subject, 1)  # Показываем только первый результат
        if urait_results:
            print(f"[DEBUG] Юрайт найден первый результат")
            result = urait_results[0]
            info["links"].append(result["url"])
            info["resources"].append(result)

            info["primary_link"] = result["url"]
            info["status"] = "success"
            info["note"] = "Найдено на Юрайт"
        else:
            print(f"[DEBUG] Юрайт не найден для: '{subject}'")
    except Exception as exc:
        print(f"[DEBUG] Ошибка поиска Юрайт: {exc}")
        info["status"] = "error"
        info["note"] = f"Ошибка urait: {exc}"

    ipr_result = fetch_iprbookshop_reader(subject)
    if ipr_result:
        # Всегда одиночный результат IPRbooks
        ipr_resource = {
            "title": ipr_result.get("title", f"Учебник по предмету: {subject}"),
            "url": ipr_result["url"],
            "status": ipr_result.get("status", "success"),
            "note": f"IPRbooks - {ipr_result.get('note', 'Электронная библиотека с полным доступом к тексту')}",
            "source": "iprbookshop"
        }
        info["links"].append(ipr_result["url"])
        info["resources"].append(ipr_resource)
        info["primary_link"] = info["primary_link"] or ipr_result["url"]
        info["status"] = ipr_result.get("status", "success")
        info["note"] = info.get("note") or "Найдено на IPRbooks"

    seen_links: Set[str] = set()
    unique_links: List[str] = []
    unique_resources: List[Dict[str, Any]] = []
    for res in info["resources"]:
        url = res.get("url")
        if not url or url in seen_links:
            continue
        seen_links.add(url)
        unique_links.append(url)
        unique_resources.append(res)

    info["links"] = unique_links
    info["resources"] = unique_resources

    if info["status"] == "error" and not unique_links:
        info["note"] = info.get("note") or "Ссылки не найдены"

    return subject, info


def copy_style(src: Cell, dst: Cell) -> None:
    if src is None or dst is None:
        return
    if not src.has_style:
        return
    try:
        val = src.font
        dst.font = shallow_copy(val)
    except Exception:
        pass
    try:
        val = src.border
        dst.border = shallow_copy(val)
    except Exception:
        pass
    try:
        val = src.fill
        dst.fill = shallow_copy(val)
    except Exception:
        pass
    try:
        dst.number_format = src.number_format
    except Exception:
        pass
    try:
        val = src.protection
        dst.protection = shallow_copy(val)
    except Exception:
        pass
    try:
        val = src.alignment
        dst.alignment = shallow_copy(val)
    except Exception:
        pass


def normalize_subject(value: str) -> str:
    return value.strip().lower()


def clean_data_for_json(data):
    """Очищает данные от проблемных символов для JSON"""
    if isinstance(data, dict):
        return {k: clean_data_for_json(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [clean_data_for_json(item) for item in data]
    elif isinstance(data, str):
        # Удаляем или заменяем проблемные символы
        cleaned = data.replace('\x00', '').replace('\r', ' ').replace('\n', ' ')
        # Удаляем Unicode BOM и другие невидимые символы
        cleaned = cleaned.replace('\ufeff', '').replace('\u200b', '').replace('\u00a0', ' ')
        # Удаляем множественные пробелы
        import re
        cleaned = re.sub(r'\s+', ' ', cleaned)
        # Удаляем управляющие символы, оставляя только печатные
        cleaned = ''.join(char for char in cleaned if ord(char) >= 32 or char in ' \t')
        # Обрезаем пробелы в начале и конце
        cleaned = cleaned.strip()
        return cleaned
    else:
        return data


@app.route("/process", methods=["POST"])
def process():
    return process_streaming()


@app.route("/get_next_resource", methods=["POST"])
def get_next_resource():
    """Получает следующий вариант ресурса для предмета"""
    data = request.get_json()
    subject = data.get("subject")
    current_url = data.get("current_url")
    source = data.get("source", "")
    
    print(f"[DEBUG] Запрос следующего ресурса для '{subject}', текущий: {current_url}")
    
    if source == "rmebrk":
        # Для RMЭБ запускаем поиск заново
        print(f"[DEBUG] RMЭБ: ищем следующий вариант для '{subject}'")
        try:
            max_results = 5 if os.getenv('VERCEL') == '1' else 10
            rmebrk_results = search_rmebrk_results(subject, max_results)
            if not rmebrk_results:
                return jsonify({"status": "no_more", "message": "Результаты не найдены"})

            # Находим текущий индекс по URL
            current_index = -1
            for i, result in enumerate(rmebrk_results):
                if current_url in result["url"] or result["url"] in current_url:
                    current_index = i
                    print(f"[DEBUG] RMЭБ: найден текущий индекс: {i}")
                    break

            next_index = current_index + 1
            if next_index < len(rmebrk_results):
                return jsonify({"status": "success", "resource": rmebrk_results[next_index]})
            else:
                return jsonify({"status": "no_more", "message": "Больше вариантов нет"})

        except Exception as e:
            print(f"[DEBUG] RMЭБ ошибка поиска следующего: {e}")
            return jsonify({"status": "no_more", "message": "Ошибка при поиске вариантов"})

    if source == "urait":
        results = search_urait_multiple_results(subject, 10)
        # Находим текущий индекс и возвращаем следующий
        current_index = -1
        for i, result in enumerate(results):
            if result["url"] == current_url:
                current_index = i
                break

        next_index = current_index + 1
        if next_index < len(results):
            return jsonify({"status": "success", "resource": results[next_index]})
        else:
            return jsonify({"status": "no_more", "message": "Больше вариантов нет"})

    elif source == "iprbookshop":
        # Для IPRbooks запускаем поиск заново через Playwright
        print(f"[DEBUG] IPRbooks: ищем следующий вариант для '{subject}'")
        try:
            ipr_results = search_iprbookshop_multiple_results(subject, 10)
            if not ipr_results:
                return jsonify({"status": "no_more", "message": "Результаты не найдены"})
            
            # Находим текущий индекс по ID публикации
            current_index = -1
            current_publication_id = None
            
            # Извлекаем ID из текущего URL (например: publicationId=149958)
            if "publicationId=" in current_url:
                current_publication_id = current_url.split("publicationId=")[1].split("&")[0]
                print(f"[DEBUG] IPRbooks: ищем по ID публикации: {current_publication_id}")
            
            for i, result in enumerate(ipr_results):
                # Проверяем по ID публикации или частичному совпадению URL
                if current_publication_id and current_publication_id in result["url"]:
                    current_index = i
                    print(f"[DEBUG] IPRbooks: найден текущий индекс по ID: {i}")
                    break
                elif current_url in result["url"] or result["url"] in current_url:
                    current_index = i
                    print(f"[DEBUG] IPRbooks: найден текущий индекс по URL: {i}")
                    break
            
            next_index = current_index + 1
            if next_index < len(ipr_results):
                return jsonify({"status": "success", "resource": ipr_results[next_index]})
            else:
                return jsonify({"status": "no_more", "message": "Больше вариантов нет"})
                
        except Exception as e:
            print(f"[DEBUG] IPRbooks ошибка поиска следующего: {e}")
            return jsonify({"status": "error", "message": f"Ошибка поиска: {e}"})
    
    return jsonify({"status": "error", "message": "Неизвестный источник"})


@app.route("/process_streaming", methods=["POST"])
def process_streaming():
    ood_file = request.files.get("ood_file")
    up_file = request.files.get("up_file")
    if not ood_file or not up_file:
        return jsonify({"status": "error", "message": "Загрузите оба файла: ООД и 33-УП."}), 400

    subjects = extract_subjects_from_up33(up_file)
    if not subjects:
        return jsonify({"status": "error", "message": "В 33-УП не найдено подходящих предметов."}), 400

    ood_file.stream.seek(0)
    wb = load_workbook(ood_file, data_only=False)
    ws = wb.active

    start_row = 136
    existing_subjects: Set[str] = set()
    existing_links: Set[Tuple[str, str]] = set()

    # Check existing subjects from ALL rows (1 to max_row) to avoid duplicates
    for r in range(1, ws.max_row + 1):
        subj_val = ws.cell(row=r, column=2).value
        if subj_val:
            existing_subjects.add(normalize_subject(str(subj_val)))

    for r in range(start_row, ws.max_row + 1):
        subj_val = ws.cell(row=r, column=2).value
        link_val = ws.cell(row=r, column=4).value
        if subj_val and link_val:
            existing_links.add((normalize_subject(str(subj_val)), str(link_val).strip()))

    pending_subjects: List[str] = []
    skipped_subjects: List[Dict[str, str]] = []
    for subject in subjects:
        norm = normalize_subject(subject)
        if not norm:
            continue
        if norm in existing_subjects:
            skipped_subjects.append({"subject": subject, "reason": "Уже заполнено в ООД"})
            continue
        pending_subjects.append(subject)
        existing_subjects.add(norm)

    if not pending_subjects:
        return jsonify({"status": "ok", "results": [], "skipped": skipped_subjects})

    def generate():
        import json
        
        # Send initial subjects list
        for subject in pending_subjects:
            cleaned_subject = clean_data_for_json(subject)
            yield f"data: {json.dumps({'type': 'subject_start', 'subject': cleaned_subject}, ensure_ascii=False, separators=(',', ':'))}\n\n"
        
        link_results: Dict[str, Dict[str, Any]] = {}
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_subject = {executor.submit(fetch_links_for_subject, subject): subject for subject in pending_subjects}
            for future in future_to_subject:
                subject_key = future_to_subject[future]
                try:
                    subject, info = future.result()
                    link_results[subject] = info
                    
                    # Clean data to prevent JSON parsing errors
                    cleaned_info = clean_data_for_json(info)
                    
                    # Escape JSON properly to avoid parse errors
                    cleaned_subject = clean_data_for_json(subject)
                    json_data = json.dumps({'type': 'subject_done', 'subject': cleaned_subject, 'info': cleaned_info}, ensure_ascii=False, separators=(',', ':'))
                    yield f"data: {json_data}\n\n"
                except Exception as exc:
                    fallback_info = {
                        "links": [],
                        "primary_link": None,
                        "status": "error",
                        "note": f"Ошибка поиска: {exc}",
                        "resources": []
                    }
                    link_results[subject_key] = fallback_info
                    cleaned_subject_key = clean_data_for_json(subject_key)
                    cleaned_fallback_info = clean_data_for_json(fallback_info)
                    json_data = json.dumps({'type': 'subject_done', 'subject': cleaned_subject_key, 'info': cleaned_fallback_info}, ensure_ascii=False, separators=(',', ':'))
                    yield f"data: {json_data}\n\n"

        # Generate final file
        results_payload = process_excel_file(ws, pending_subjects, link_results, start_row)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Отправляем результаты частями, чтобы избежать слишком больших JSON
        chunk_size = 5  # По 5 результатов за раз
        results_chunks = [results_payload[i:i + chunk_size] for i in range(0, len(results_payload), chunk_size)]
        
        for i, chunk in enumerate(results_chunks):
            chunk_data = {
                "type": "results_chunk",
                "chunk_index": i,
                "total_chunks": len(results_chunks),
                "results": clean_data_for_json(chunk)
            }
            yield f"data: {json.dumps(chunk_data, ensure_ascii=False, separators=(',', ':'))}\n\n"
        
        # Отправляем финальные данные без больших массивов
        final_data = {
            "type": "complete",
            "status": "ok",
            "total_results": len(results_payload),
            "skipped": clean_data_for_json(skipped_subjects),
            "file_data": base64.b64encode(output.getvalue()).decode("utf-8"),
            "filename": "updated_ood.xlsx",
        }
        yield f"data: {json.dumps(final_data, ensure_ascii=False, separators=(',', ':'))}\n\n"
    
    from flask import Response
    return Response(generate(), mimetype='text/event-stream')


def process_excel_file(ws, pending_subjects, link_results, start_row):
    row_ptr = find_next_row(ws, start_row)
    next_num = compute_next_number(ws, start_row)

    template_a = ws.cell(row=start_row, column=1)
    template_b = ws.cell(row=start_row, column=2)
    template_d = ws.cell(row=start_row, column=4)

    results_payload: List[Dict[str, Any]] = []
    existing_links: Set[Tuple[str, str]] = set()
    
    # Get existing links from rows 136+
    for r in range(start_row, ws.max_row + 1):
        subj_val = ws.cell(row=r, column=2).value
        link_val = ws.cell(row=r, column=4).value
        if subj_val and link_val:
            existing_links.add((normalize_subject(str(subj_val)), str(link_val).strip()))
    
    for subject in pending_subjects:
        row_ptr = find_next_row(ws, row_ptr)
        a_cell = ws.cell(row=row_ptr, column=1)
        b_cell = ws.cell(row=row_ptr, column=2)
        copy_style(template_a, a_cell)
        copy_style(template_b, b_cell)
        a_cell.value = next_num
        b_cell.value = subject

        next_num += 1
        row_ptr += 1

        subject_info = link_results.get(subject, {"links": []})
        links = subject_info.get("links", [])
        primary_link = subject_info.get("primary_link")
        status = subject_info.get("status", "warning")
        note = subject_info.get("note", "")

        norm_subject = normalize_subject(subject)
        for link in links:
            key = (norm_subject, link)
            if key in existing_links:
                continue
            row_ptr = find_next_row(ws, row_ptr)
            b = ws.cell(row=row_ptr, column=2)
            d = ws.cell(row=row_ptr, column=4)
            copy_style(template_b, b)
            copy_style(template_d, d)
            b.value = subject
            d.value = link
            existing_links.add(key)
            row_ptr += 1

        if status == "error" and not links:
            note = note or "Ссылки не найдены"
        elif status == "warning" and not note:
            note = "Требуется выбрать книгу вручную"

        results_payload.append(
            {
                "subject": subject,
                "status": status,
                "primary_link": primary_link,
                "links": links,
                "resources": subject_info.get("resources", []),
                "note": note,
            }
        )

    return results_payload


if __name__ == "__main__":
    app.run(debug=True)
